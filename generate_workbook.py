# Event-Inventory Command Center Workbook Generator
# This Python script generates the complete Excel workbook with all sheets, tables, and formulas

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.comments import Comment
import os

def create_command_center_workbook():
    """Generate the complete Event-Inventory-CommandCenter.xlsx workbook"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    warning_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    danger_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    # Sheets to create (ordered)
    sheets_config = [
        ("Dashboard", create_dashboard_sheet),
        ("Plan_Buy", create_plan_buy_sheet),
        ("ROP_SS", create_rop_ss_sheet),
        ("PnL", create_pnl_sheet),
        ("Shrink", create_shrink_sheet),
        ("Counts_Entry", create_counts_entry_sheet),
        ("Forms_Inbox", create_forms_inbox_sheet),
        ("Counts_Manual", create_counts_manual_sheet),
        ("Events", create_events_sheet),
        ("SKU", create_sku_sheet),
        ("Sales", create_sales_sheet),
        ("Audits", create_audits_sheet),
        ("Config", create_config_sheet),
        ("Mappings", create_mappings_sheet),
        ("Calendar", create_calendar_sheet),
        ("Staging", create_staging_sheet),
        ("Exceptions", create_exceptions_sheet),
    ]
    
    for sheet_name, creator_func in sheets_config:
        ws = wb.create_sheet(sheet_name)
        creator_func(ws, header_fill, header_font)
        
        # Apply table formatting if this is a data sheet
        if sheet_name not in ["Dashboard", "Staging"]:
            format_as_table(ws, sheet_name)
    
    # Set Dashboard as active sheet
    wb.active = wb["Dashboard"]
    
    return wb

def create_dashboard_sheet(ws, header_fill, header_font):
    """Create the Dashboard sheet with KPIs and visuals"""
    
    # Local styling
    warning_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    # Title
    ws["A1"] = "Event Inventory Command Center"
    ws["A1"].font = Font(size=18, bold=True, color="366092")
    ws.merge_cells("A1:H1")
    
    # Last refresh timestamp
    ws["A2"] = "Last Refresh:"
    ws["B2"] = "=NOW()"
    ws["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    
    # KPI Section
    ws["A4"] = "Key Performance Indicators"
    ws["A4"].font = Font(size=14, bold=True)
    
    # KPI Headers
    kpis = [
        ("A6", "Metric", "B6", "Current", "C6", "Target", "D6", "Status"),
        ("A7", "Total Revenue", "B7", "=SUMIF(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!E:E)", "C7", "100000", "D7", "=IF(B7>=C7,\"✓\",\"⚠\")"),
        ("A8", "Gross Margin %", "B8", "=SUMIF(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!F:F)/SUMIF(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!E:E)", "C8", "0.35", "D8", "=IF(B8>=C8,\"✓\",\"⚠\")"),
        ("A9", "GMROI", "B9", "=SUMIF(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!G:G)/COUNTIFS(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!G:G,\">0\")", "C9", "3.0", "D9", "=IF(B9>=C9,\"✓\",\"⚠\")"),
        ("A10", "Sell-Through %", "B10", "=SUMIF(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!H:H)/COUNTIFS(PnL!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),PnL!H:H,\">0\")", "C10", "0.75", "D10", "=IF(B10>=C10,\"✓\",\"⚠\")"),
        ("A11", "Days of Supply", "B11", "=AVERAGE(Plan_Buy!G:G)", "C11", "30", "D11", "=IF(AND(B11>=C11*0.8,B11<=C11*1.2),\"✓\",\"⚠\")"),
        ("A12", "Stockout Risk SKUs", "B12", "=COUNTIFS(ROP_SS!H:H,\"<0\")", "C12", "0", "D12", "=IF(B12<=C12,\"✓\",\"⚠\")"),
        ("A13", "Shrink %", "B13", "=ABS(SUMIF(Shrink!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Shrink!F:F))/SUMIF(Shrink!A:A,\">\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Shrink!D:D)", "C13", "0.02", "D13", "=IF(B13<=C13,\"✓\",\"⚠\")"),
    ]
    
    for row_data in kpis:
        for i in range(0, len(row_data), 2):
            if i + 1 < len(row_data):
                ws[row_data[i]] = row_data[i + 1]
    
    # Format KPI headers
    for col in ["A6", "B6", "C6", "D6"]:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal="center")
    
    # Event Calendar Preview
    ws["F4"] = "Upcoming Events (Next 30 Days)"
    ws["F4"].font = Font(size=12, bold=True)
    
    event_headers = ["Event", "Date", "Type", "Attendance"]
    for i, header in enumerate(event_headers):
        cell = ws.cell(row=6, column=6+i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Top Movers section
    ws["A15"] = "Top Moving SKUs (Last 30 Days)"
    ws["A15"].font = Font(size=12, bold=True)
    
    mover_headers = ["SKU", "Description", "Units Sold", "Revenue", "GM%"]
    for i, header in enumerate(mover_headers):
        cell = ws.cell(row=17, column=1+i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Risk Watchlist
    ws["G15"] = "Risk Watchlist"
    ws["G15"].font = Font(size=12, bold=True)
    
    risk_headers = ["SKU", "Risk Type", "Current", "Threshold"]
    for i, header in enumerate(risk_headers):
        cell = ws.cell(row=17, column=7+i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add conditional formatting for status column
    ws.conditional_formatting.add("D7:D13", 
        CellIsRule(operator="equal", formula=["✓"], fill=success_fill))
    ws.conditional_formatting.add("D7:D13", 
        CellIsRule(operator="equal", formula=["⚠"], fill=warning_fill))

def create_plan_buy_sheet(ws, header_fill, header_font):
    """Create Plan_Buy sheet with forecast and purchase recommendations"""
    
    headers = [
        "SKU", "Description", "Category", "Current_Stock", "Forecast_30d", 
        "Safety_Stock", "Days_of_Supply", "ROP", "Recommended_Order_Qty", 
        "Order_Cost", "GM_Impact", "Priority", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample formula rows
    sample_formulas = [
        ("A2", "SKU001", "B2", "=VLOOKUP(A2,SKU!A:B,2,FALSE)", "C2", "=VLOOKUP(A2,SKU!A:C,3,FALSE)"),
        ("D2", "=SUMIFS(Counts_Entry!E:E,Counts_Entry!C:C,\"in_store\",Counts_Entry!D:D,A2)+SUMIFS(Counts_Entry!E:E,Counts_Entry!C:C,\"back_of_store\",Counts_Entry!D:D,A2)"),
        ("E2", "=VLOOKUP(A2,ROP_SS!A:C,3,FALSE)*30"),
        ("F2", "=VLOOKUP(A2,ROP_SS!A:E,5,FALSE)"),
        ("G2", "=IF(E2>0,D2/E2*30,999)"),
        ("H2", "=VLOOKUP(A2,ROP_SS!A:F,6,FALSE)"),
        ("I2", "=MAX(0,Config!$B$3*E2+F2-D2)"),
        ("J2", "=I2*VLOOKUP(A2,SKU!A:D,4,FALSE)"),
        ("K2", "=I2*VLOOKUP(A2,SKU!A:E,5,FALSE)-J2"),
        ("L2", "=IF(D2<H2,\"HIGH\",IF(G2<Config!$B$3*0.5,\"MEDIUM\",\"LOW\"))"),
    ]
    
    for formula_set in sample_formulas:
        for i in range(0, len(formula_set), 2):
            if i + 1 < len(formula_set):
                ws[formula_set[i]] = formula_set[i + 1]

def create_rop_ss_sheet(ws, header_fill, header_font):
    """Create ROP & Safety Stock calculation sheet"""
    
    headers = [
        "SKU", "Description", "Avg_Daily_Demand", "Lead_Time_Days", 
        "Demand_StdDev", "Safety_Stock", "ROP", "Stock_Available", 
        "Days_Until_Stockout", "Action_Required"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample calculations
    ws["A2"] = "SKU001"
    ws["B2"] = "=VLOOKUP(A2,SKU!A:B,2,FALSE)"
    ws["C2"] = "=IFERROR(AVERAGEIFS(Sales!C:C,Sales!B:B,A2,Sales!A:A,\">\"&TODAY()-90)/90,0)"
    ws["D2"] = "=VLOOKUP(A2,SKU!A:F,6,FALSE)"
    ws["E2"] = "=IFERROR(STDEV.S(IF(Sales!B:B=A2,IF(Sales!A:A>TODAY()-90,Sales!C:C))),C2*0.5)"
    ws["F2"] = "=Config!$B$1*SQRT(E2^2*D2+C2^2*1)"  # Assuming LT variance = 1
    ws["G2"] = "=C2*D2+F2"
    ws["H2"] = "=SUMIFS(Counts_Entry!E:E,Counts_Entry!D:D,A2)"
    ws["I2"] = "=IF(C2>0,H2/C2,999)"
    ws["J2"] = "=IF(H2<G2,\"REORDER NOW\",IF(I2<Config!$B$3,\"MONITOR\",\"OK\"))"

def create_pnl_sheet(ws, header_fill, header_font):
    """Create P&L analysis sheet"""
    
    headers = [
        "Date", "SKU", "Description", "Category", "Revenue", "COGS", 
        "Gross_Margin", "GMROI", "Sell_Through", "Units_Sold", "Avg_Unit_Price"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font

def create_shrink_sheet(ws, header_fill, header_font):
    """Create shrink analysis sheet"""
    
    headers = [
        "Date", "Checkpoint", "Location", "SKU", "Expected_Qty", "Counted_Qty", 
        "Variance", "Shrink_Pct", "Shrink_Value", "Category", "Counter_ID", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font

def create_counts_entry_sheet(ws, header_fill, header_font):
    """Create unified counts entry sheet (canonical table)"""
    
    headers = [
        "Unique_Key", "AsOf_Date", "Checkpoint", "Location", "SKU", 
        "Qty", "UOM", "Counter_ID", "Notes", "Source", "Submitted_At", 
        "Photo_URL", "Validated"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add sample validation formulas
    ws["A2"] = "=TEXT(B2,\"yyyymmdd\")&\"|\"&C2&\"|\"&D2&\"|\"&E2&\"|\"&H2"
    ws["M2"] = "=AND(B2<>\"\",C2<>\"\",D2<>\"\",E2<>\"\",ISNUMBER(F2),F2>=0)"

def create_forms_inbox_sheet(ws, header_fill, header_font):
    """Create Forms inbox for Microsoft Forms integration"""
    
    headers = [
        "Response_ID", "Submitted_At", "Checkpoint", "Location", "Date_Counted", 
        "SKU", "Qty", "UOM", "Counter_ID", "Notes", "Photo", "Processed"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font

def create_counts_manual_sheet(ws, header_fill, header_font):
    """Create manual transcription sheet"""
    
    headers = [
        "Entry_ID", "Date_Counted", "Checkpoint", "Location", "SKU", 
        "Description", "Qty", "UOM", "Counter_ID", "Notes", "Status", "Validated"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data validation for key fields
    ws["C2"].comment = Comment("BOM, MID, or EOM", "System")
    ws["D2"].comment = Comment("in_store or back_of_store", "System")

def create_events_sheet(ws, header_fill, header_font):
    """Create events master sheet"""
    
    headers = [
        "Event_ID", "Name", "Venue_Area", "Event_Type", "Start_DateTime", 
        "End_DateTime", "Est_Attendance", "Actual_Attendance", "Status", 
        "Conversion_Rate", "Attach_Rate", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample events
    sample_events = [
        ["EVT001", "Tech Conference 2025", "Main Hall", "Conference", "2025-09-15 09:00", "2025-09-15 18:00", 500, "", "Planned", "", "", ""],
        ["EVT002", "Product Launch", "Booth A", "Launch", "2025-09-20 10:00", "2025-09-20 16:00", 250, "", "Planned", "", "", ""],
        ["EVT003", "Training Workshop", "Room B", "Training", "2025-09-25 08:00", "2025-09-25 17:00", 100, "", "Planned", "", "", ""],
    ]
    
    for i, event in enumerate(sample_events, 2):
        for j, value in enumerate(event, 1):
            ws.cell(row=i, column=j, value=value)

def create_sku_sheet(ws, header_fill, header_font):
    """Create SKU master sheet"""
    
    headers = [
        "SKU", "Description", "Category", "Cost", "Price", "Lead_Time_Days", 
        "Supplier", "UOM", "Active", "Min_Order_Qty", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample SKUs
    sample_skus = [
        ["SKU001", "UPS Branded Pen", "Promotional", 0.50, 2.00, 14, "PromoCorp", "EA", "Y", 100, ""],
        ["SKU002", "Shipping Box Small", "Packaging", 1.25, 3.50, 7, "BoxCo", "EA", "Y", 50, ""],
        ["SKU003", "UPS T-Shirt L", "Apparel", 8.00, 25.00, 21, "ApparelPlus", "EA", "Y", 25, ""],
        ["SKU004", "Bubble Wrap Roll", "Packaging", 15.00, 35.00, 10, "PackingSupply", "ROLL", "Y", 10, ""],
        ["SKU005", "UPS Coffee Mug", "Promotional", 3.25, 12.00, 14, "PromoCorp", "EA", "Y", 48, ""],
    ]
    
    for i, sku in enumerate(sample_skus, 2):
        for j, value in enumerate(sku, 1):
            ws.cell(row=i, column=j, value=value)

def create_sales_sheet(ws, header_fill, header_font):
    """Create sales history sheet"""
    
    headers = [
        "Date", "SKU", "Units_Sold", "Revenue", "Event_ID", "Channel", 
        "Salesperson", "Discount_Pct", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font

def create_audits_sheet(ws, header_fill, header_font):
    """Create historical audits archive"""
    
    headers = [
        "Audit_ID", "Date", "Checkpoint", "Location", "SKU", "Qty", 
        "UOM", "Counter_ID", "Auditor", "Notes", "Archived_At"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font

def create_config_sheet(ws, header_fill, header_font):
    """Create configuration parameters sheet"""
    
    # Configuration table
    config_items = [
        ("Z Service Level", "B1", 1.65, "Standard deviations for safety stock"),
        ("Default Lead Time Days", "B2", 14, "When SKU lead time is missing"),
        ("Target Days of Supply", "B3", 30, "Inventory planning horizon"),
        ("Max Cash Per Order", "B4", 50000, "Budget constraint per purchase order"),
        ("Shrink Threshold %", "B5", 0.05, "Flag shrink above this percentage"),
        ("Low DoS Warning", "B6", 15, "Days of supply warning threshold"),
        ("Default Event Conversion", "B7", 0.15, "Default event conversion rate"),
        ("Default Attach Rate", "B8", 1.2, "Default items per transaction"),
        ("Forms Integration Enabled", "B9", "TRUE", "Enable Microsoft Forms intake"),
        ("Manual Entry Enabled", "B10", "TRUE", "Enable manual transcription"),
    ]
    
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["C1"] = "Description"
    
    for i, (param, cell, value, desc) in enumerate(config_items):
        ws[f"A{i+2}"] = param
        ws[cell] = value
        ws[f"C{i+2}"] = desc
    
    # Format header
    for col in ["A1", "B1", "C1"]:
        ws[col].fill = header_fill
        ws[col].font = header_font

def create_mappings_sheet(ws, header_fill, header_font):
    """Create event mappings for conversion and attach rates"""
    
    # Conversion rates by event type and category
    ws["A1"] = "Conversion Rates"
    ws["A1"].font = Font(size=12, bold=True)
    
    conv_headers = ["Event_Type", "Category", "Conversion_Rate"]
    for i, header in enumerate(conv_headers, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample conversion rates
    conv_data = [
        ["Conference", "Promotional", 0.20],
        ["Conference", "Apparel", 0.10],
        ["Launch", "Promotional", 0.35],
        ["Launch", "Packaging", 0.05],
        ["Training", "Promotional", 0.15],
    ]
    
    for i, row_data in enumerate(conv_data, 4):
        for j, value in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=value)
    
    # Attach rates
    ws["E1"] = "Attach Rates"
    ws["E1"].font = Font(size=12, bold=True)
    
    attach_headers = ["Event_Type", "Category", "Attach_Rate"]
    for i, header in enumerate(attach_headers):
        cell = ws.cell(row=3, column=5+i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample attach rates
    attach_data = [
        ["Conference", "Promotional", 1.5],
        ["Conference", "Apparel", 1.0],
        ["Launch", "Promotional", 2.0],
        ["Launch", "Packaging", 1.2],
        ["Training", "Promotional", 1.3],
    ]
    
    for i, row_data in enumerate(attach_data, 4):
        for j, value in enumerate(row_data):
            ws.cell(row=i, column=5+j, value=value)

def create_calendar_sheet(ws, header_fill, header_font):
    """Create calendar spine sheet"""
    
    headers = [
        "Date", "Day_of_Week", "Is_Weekend", "Is_Holiday", "Holiday_Name", 
        "Event_Window", "Event_Count", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Generate 365 days starting from today
    base_date = datetime.now().date()
    for i in range(365):
        current_date = base_date + timedelta(days=i)
        row = i + 2
        
        ws[f"A{row}"] = current_date
        ws[f"B{row}"] = current_date.strftime("%A")
        ws[f"C{row}"] = "TRUE" if current_date.weekday() >= 5 else "FALSE"
        # Add formulas for event counting
        ws[f"G{row}"] = f"=COUNTIFS(Events!E:E,\">=\"&A{row},Events!F:F,\"<=\"&A{row}+1)"

def create_staging_sheet(ws, header_fill, header_font):
    """Create staging area for Power Query transformations"""
    
    ws["A1"] = "STAGING AREA - DO NOT EDIT DIRECTLY"
    ws["A1"].font = Font(size=14, bold=True, color="FF0000")
    ws["A3"] = "This sheet is used by Power Query for data transformations."
    ws["A4"] = "Data will appear here during refresh operations."
    ws["A5"] = "Sheet may be hidden in production."

def create_exceptions_sheet(ws, header_fill, header_font):
    """Create exceptions tracking sheet"""
    
    headers = [
        "Exception_ID", "Date_Found", "Source_Sheet", "Source_Row", "Exception_Type", 
        "Description", "Severity", "Resolution_Required", "Status", "Notes"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample exception types
    sample_exceptions = [
        ["EXC001", "2025-08-20", "Forms_Inbox", "5", "Invalid SKU", "SKU 'XYZ999' not found in master", "HIGH", "Verify and correct SKU", "OPEN", "Check with counter"],
        ["EXC002", "2025-08-20", "Counts_Manual", "12", "Duplicate Entry", "Same count submitted twice", "MEDIUM", "Remove duplicate", "RESOLVED", "Kept latest entry"],
        ["EXC003", "2025-08-20", "Counts_Entry", "25", "Negative Quantity", "Qty = -5 for SKU001", "HIGH", "Verify count", "OPEN", "Impossible negative stock"],
    ]
    
    for i, exc in enumerate(sample_exceptions, 2):
        for j, value in enumerate(exc, 1):
            ws.cell(row=i, column=j, value=value)

def format_as_table(ws, sheet_name):
    """Apply table formatting and create named range"""
    
    # Find the data range
    max_row = ws.max_row if ws.max_row > 1 else 2
    max_col = ws.max_column
    
    # Create table reference
    table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    
    # Apply borders and formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[table_ref]:
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:  # Header row
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Freeze top row
    ws.freeze_panes = "A2"

def main():
    """Main function to generate the workbook"""
    print("Generating Event-Inventory-CommandCenter.xlsx...")
    
    wb = create_command_center_workbook()
    
    # Save workbook
    filename = "Event-Inventory-CommandCenter.xlsx"
    wb.save(filename)
    
    print(f"✅ Workbook '{filename}' created successfully!")
    print(f"📊 Contains {len(wb.sheetnames)} worksheets:")
    for sheet in wb.sheetnames:
        print(f"   • {sheet}")
    
    print("\n🚀 Next Steps:")
    print("1. Open the workbook in Excel")
    print("2. Load your event and SKU data")
    print("3. Set up Microsoft Forms integration")
    print("4. Start collecting inventory counts")
    print("5. Run forecasts and generate buy plans")

if __name__ == "__main__":
    main()
